from io import BytesIO
from pathlib import Path
import tempfile

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from parser_pdf_base_mestre_v1_3_estavel import (
    processar_pdf,
    gerar_resumo,
    validar_estrutura,
)

st.set_page_config(page_title="Sistema de Cobrança - Módulo 1", page_icon="📄", layout="wide")
st.title("Sistema de Cobrança — Módulo 1")
st.subheader("PDF → Base Mestre")
st.write("Selecione os relatórios de inadimplência em PDF. Somente PDFs validados entram na Base Mestre.")

def resultado_esta_ok(resultado):
    validacao = resultado.get("validacao", {})
    return bool(validacao) and all(item.get("ok", False) for item in validacao.values())

def gerar_csv_bytes(df):
    return df.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")

def gerar_excel_bytes(base_mestre, resumo, validacao_estrutura):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        base_mestre.to_excel(writer, sheet_name="Dados", index=False)
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        validacao_estrutura.to_excel(writer, sheet_name="Validação", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    cor = "4F81BD"
    borda = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for nome_aba in ["Dados", "Resumo", "Validação"]:
        ws = wb[nome_aba]
        for linha in ws.iter_rows():
            for cell in linha:
                cell.border = borda

        for cell in ws[1]:
            cell.fill = PatternFill(fill_type="solid", fgColor=cor)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for coluna in ws.columns:
            letra = get_column_letter(coluna[0].column)
            maior = max(len("" if c.value is None else str(c.value)) for c in coluna)
            ws.column_dimensions[letra].width = min(maior + 2, 45)

    ws = wb["Dados"]
    cab = {c.value: c.column for c in ws[1]}
    for nome in ["vlr_base", "multa", "juros", "correcao", "vlr_corrigido"]:
        if nome in cab:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=cab[nome]).number_format = '#,##0.00'

    ws = wb["Resumo"]
    cab = {c.value: c.column for c in ws[1]}
    for nome in ["valor_total_devido_base", "valor_total_corrigido"]:
        if nome in cab:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=cab[nome]).number_format = '#,##0.00'

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def processar_uploads(arquivos):
    bases = []
    relatorio = []

    with tempfile.TemporaryDirectory(prefix="cobranca_") as tmp:
        pasta = Path(tmp)
        for arq in arquivos:
            caminho = pasta / arq.name
            caminho.write_bytes(arq.getbuffer())

            try:
                resultado = processar_pdf(caminho)
                valido = resultado_esta_ok(resultado)
                dados = resultado["dados"]
                cab = resultado.get("cabecalho", {})

                relatorio.append({
                    "arquivo": arq.name,
                    "codigo_condominio": cab.get("codigo_condominio", ""),
                    "condominio": cab.get("condominio", ""),
                    "registros": len(dados),
                    "status": "OK" if valido else "DIVERGENTE",
                })

                if valido:
                    bases.append(dados)
            except Exception as e:
                relatorio.append({
                    "arquivo": arq.name,
                    "codigo_condominio": "",
                    "condominio": "",
                    "registros": 0,
                    "status": f"ERRO: {e}",
                })

    base = pd.concat(bases, ignore_index=True) if bases else pd.DataFrame()
    return base, pd.DataFrame(relatorio)

arquivos = st.file_uploader(
    "Selecione um ou mais PDFs",
    type=["pdf"],
    accept_multiple_files=True,
)

if st.button("Processar arquivos", type="primary", disabled=not arquivos):
    with st.spinner("Processando..."):
        base, relatorio = processar_uploads(arquivos)
    st.session_state["base"] = base
    st.session_state["relatorio"] = relatorio

if "relatorio" in st.session_state:
    relatorio = st.session_state["relatorio"]
    base = st.session_state["base"]

    c1, c2, c3, c4 = st.columns(4)
    validos = int((relatorio["status"] == "OK").sum())
    c1.metric("PDFs processados", len(relatorio))
    c2.metric("PDFs validados", validos)
    c3.metric("Com erro/divergência", len(relatorio) - validos)
    c4.metric("Registros consolidados", len(base))

    st.subheader("Relatório de processamento")
    st.dataframe(relatorio, width="stretch", hide_index=True)

    if not base.empty:
        resumo = gerar_resumo(base)
        validacao = validar_estrutura(base)

        st.subheader("Resumo")
        st.dataframe(resumo, width="stretch", hide_index=True)

        if len(validacao) == 0:
            st.success("Nenhum alerta estrutural encontrado.")
        else:
            st.warning(f"{len(validacao)} alerta(s) estrutural(is). Consulte a aba Validação do Excel.")

        csv_bytes = gerar_csv_bytes(base)
        xlsx_bytes = gerar_excel_bytes(base, resumo, validacao)

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "Baixar Base Mestre (CSV)",
                data=csv_bytes,
                file_name="base_mestre.csv",
                mime="text/csv",
                width="stretch",
            )
        with col2:
            st.download_button(
                "Baixar Base Mestre (Excel)",
                data=xlsx_bytes,
                file_name="base_mestre.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
    else:
        st.error("Nenhum PDF foi validado.")
