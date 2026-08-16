# =============================================================================
# SISTEMA DE COBRANÇA - MÓDULO 1
# Interface Desktop
#
# Esta interface utiliza o parser estável:
# parser_pdf_base_mestre_v1_3_estavel.py
#
# Não altera nenhuma regra do parser.
# =============================================================================

from pathlib import Path
import tkinter as tk
from tkinter import (
    filedialog,
    messagebox,
    ttk
)

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter

from parser_pdf_base_mestre_v1_3_estavel import (
    processar_pdf,
    gerar_resumo,
    validar_estrutura
)


# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

def resultado_esta_ok(resultado):
    """
    Retorna True somente se todas as validações
    financeiras do PDF estiverem OK.
    """

    validacao = resultado.get(
        "validacao",
        {}
    )

    if not validacao:
        return False

    return all(
        item.get("ok", False)
        for item in validacao.values()
    )


def formatar_excel(arquivo_excel):
    """
    Aplica a formatação padrão ao Excel gerado.
    """

    wb = load_workbook(
        arquivo_excel
    )

    cor_cabecalho = "4F81BD"

    borda_fina = Border(
        left=Side(
            style="thin",
            color="B7B7B7"
        ),
        right=Side(
            style="thin",
            color="B7B7B7"
        ),
        top=Side(
            style="thin",
            color="B7B7B7"
        ),
        bottom=Side(
            style="thin",
            color="B7B7B7"
        )
    )

    for nome_aba in [
        "Dados",
        "Resumo",
        "Validação"
    ]:

        ws = wb[nome_aba]

        # -----------------------------------------------------
        # Bordas
        # -----------------------------------------------------

        for linha in ws.iter_rows():
            for cell in linha:
                cell.border = borda_fina

        # -----------------------------------------------------
        # Cabeçalho
        # -----------------------------------------------------

        for cell in ws[1]:

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=cor_cabecalho
            )

            cell.font = Font(
                color="FFFFFF",
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # -----------------------------------------------------
        # Congelar cabeçalho
        # -----------------------------------------------------

        ws.freeze_panes = "A2"

        # -----------------------------------------------------
        # Filtro
        # -----------------------------------------------------

        ws.auto_filter.ref = ws.dimensions

        # -----------------------------------------------------
        # Ajustar largura
        # -----------------------------------------------------

        for coluna in ws.columns:

            letra = get_column_letter(
                coluna[0].column
            )

            maior = 0

            for cell in coluna:

                valor = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                maior = max(
                    maior,
                    len(valor)
                )

            ws.column_dimensions[
                letra
            ].width = min(
                maior + 2,
                45
            )

    # =========================================================================
    # Formatação monetária - aba Dados
    # =========================================================================

    ws_dados = wb["Dados"]

    cabecalhos_dados = {
        cell.value: cell.column
        for cell in ws_dados[1]
    }

    for nome_coluna in [
        "vlr_base",
        "multa",
        "juros",
        "correcao",
        "vlr_corrigido"
    ]:

        if nome_coluna in cabecalhos_dados:

            coluna = cabecalhos_dados[
                nome_coluna
            ]

            for linha in range(
                2,
                ws_dados.max_row + 1
            ):

                ws_dados.cell(
                    row=linha,
                    column=coluna
                ).number_format = '#,##0.00'

    # =========================================================================
    # Formatação monetária - aba Resumo
    # =========================================================================

    ws_resumo = wb["Resumo"]

    cabecalhos_resumo = {
        cell.value: cell.column
        for cell in ws_resumo[1]
    }

    for nome_coluna in [
        "valor_total_devido_base",
        "valor_total_corrigido"
    ]:

        if nome_coluna in cabecalhos_resumo:

            coluna = cabecalhos_resumo[
                nome_coluna
            ]

            for linha in range(
                2,
                ws_resumo.max_row + 1
            ):

                ws_resumo.cell(
                    row=linha,
                    column=coluna
                ).number_format = '#,##0.00'

    wb.save(
        arquivo_excel
    )


# =============================================================================
# APLICAÇÃO
# =============================================================================

class SistemaCobrancaApp:

    def __init__(self, root):

        self.root = root

        self.root.title(
            "Sistema de Cobrança"
        )

        self.root.geometry(
            "1100x850"
        )

        self.root.minsize(
            1000,
            780
        )

        # =====================================================
        # ESTILO
        # =====================================================

        style = ttk.Style()

        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(
            "Titulo.TLabel",
            font=("Arial", 22, "bold")
        )

        style.configure(
            "Subtitulo.TLabel",
            font=("Arial", 11)
        )

        style.configure(
            "Indicador.TLabel",
            font=("Arial", 11, "bold")
        )

        style.configure(
            "Status.TLabel",
            font=("Arial", 10, "bold")
        )

        style.configure(
            "Acao.TButton",
            font=("Arial", 11, "bold"),
            padding=10
        )

        style.configure(
            "Secundario.TButton",
            font=("Arial", 10),
            padding=8
        )

        # =====================================================
        # DADOS DA EXECUÇÃO
        # =====================================================

        self.arquivos_pdf = []

        self.base_mestre = None

        self.resumo = None

        self.validacao_estrutura = None

        self.relatorio = None

        # =====================================================
        # CABEÇALHO
        # =====================================================

        frame_cabecalho = ttk.Frame(
            root,
            padding=(20, 18, 20, 8)
        )

        frame_cabecalho.pack(
            fill="x"
        )

        titulo = ttk.Label(
            frame_cabecalho,
            text="Sistema de Cobrança",
            style="Titulo.TLabel"
        )

        titulo.pack()

        subtitulo = ttk.Label(
            frame_cabecalho,
            text=(
                "Módulo 1 — Conversão de relatórios "
                "de inadimplência em Base Mestre"
            ),
            style="Subtitulo.TLabel"
        )

        subtitulo.pack(
            pady=(4, 0)
        )

        # =====================================================
        # SELEÇÃO DE ARQUIVOS
        # =====================================================

        frame_selecao = ttk.LabelFrame(
            root,
            text="1. Seleção dos relatórios",
            padding=15
        )

        frame_selecao.pack(
            fill="x",
            padx=20,
            pady=(5, 10)
        )

        frame_botoes = ttk.Frame(
            frame_selecao
        )

        frame_botoes.pack()

        self.botao_selecionar = ttk.Button(
            frame_botoes,
            text="Selecionar PDFs",
            command=self.selecionar_pdfs,
            style="Acao.TButton"
        )

        self.botao_selecionar.grid(
            row=0,
            column=0,
            padx=6
        )

        self.botao_processar = ttk.Button(
            frame_botoes,
            text="Processar arquivos",
            command=self.processar_arquivos,
            state="disabled",
            style="Acao.TButton"
        )

        self.botao_processar.grid(
            row=0,
            column=1,
            padx=6
        )

        self.label_selecionados = ttk.Label(
            frame_selecao,
            text="Nenhum PDF selecionado."
        )

        self.label_selecionados.pack(
            pady=(10, 4)
        )

        # =====================================================
        # LISTA DOS ARQUIVOS SELECIONADOS
        # =====================================================

        frame_lista = ttk.Frame(
            frame_selecao
        )

        frame_lista.pack(
            fill="both",
            expand=True,
            padx=5,
            pady=(5, 0)
        )

        self.lista_arquivos = ttk.Treeview(
            frame_lista,
            columns=("arquivo",),
            show="headings",
            height=6,
            selectmode="browse"
        )

        self.lista_arquivos.heading(
            "arquivo",
            text="Arquivos selecionados"
        )

        self.lista_arquivos.column(
            "arquivo",
            anchor="w",
            width=700
        )

        scroll_lista = ttk.Scrollbar(
            frame_lista,
            orient="vertical",
            command=self.lista_arquivos.yview
        )

        self.lista_arquivos.configure(
            yscrollcommand=scroll_lista.set
        )

        self.lista_arquivos.pack(
            side="left",
            fill="both",
            expand=True
        )

        scroll_lista.pack(
            side="right",
            fill="y"
        )

        # =====================================================
        # INDICADORES
        # =====================================================

        frame_indicadores = ttk.LabelFrame(
            root,
            text="2. Resultado do processamento",
            padding=12
        )

        frame_indicadores.pack(
            fill="x",
            padx=20,
            pady=(0, 10)
        )

        indicadores = ttk.Frame(
            frame_indicadores
        )

        indicadores.pack()

        self.label_processados = ttk.Label(
            indicadores,
            text="PDFs processados: 0",
            style="Indicador.TLabel"
        )

        self.label_processados.grid(
            row=0,
            column=0,
            padx=20
        )

        self.label_validos = ttk.Label(
            indicadores,
            text="PDFs validados: 0",
            style="Indicador.TLabel"
        )

        self.label_validos.grid(
            row=0,
            column=1,
            padx=20
        )

        self.label_divergentes = ttk.Label(
            indicadores,
            text="Arquivos com problema: 0",
            style="Indicador.TLabel"
        )

        self.label_divergentes.grid(
            row=0,
            column=2,
            padx=20
        )

        self.label_registros = ttk.Label(
            indicadores,
            text="Registros consolidados: 0",
            style="Indicador.TLabel"
        )

        self.label_registros.grid(
            row=0,
            column=3,
            padx=20
        )

        # =====================================================
        # TABELA DE RESULTADOS
        # =====================================================

        frame_tabela = ttk.Frame(
            root
        )

        frame_tabela.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=(0, 10)
        )

        colunas = (
            "arquivo",
            "codigo",
            "condominio",
            "registros",
            "status"
        )

        self.tabela = ttk.Treeview(
            frame_tabela,
            columns=colunas,
            show="headings",
            height=10
        )

        self.tabela.heading(
            "arquivo",
            text="Arquivo"
        )

        self.tabela.heading(
            "codigo",
            text="Código"
        )

        self.tabela.heading(
            "condominio",
            text="Condomínio"
        )

        self.tabela.heading(
            "registros",
            text="Registros"
        )

        self.tabela.heading(
            "status",
            text="Situação"
        )

        self.tabela.column(
            "arquivo",
            width=160
        )

        self.tabela.column(
            "codigo",
            width=90,
            anchor="center"
        )

        self.tabela.column(
            "condominio",
            width=380
        )

        self.tabela.column(
            "registros",
            width=100,
            anchor="center"
        )

        self.tabela.column(
            "status",
            width=130,
            anchor="center"
        )

        scrollbar = ttk.Scrollbar(
            frame_tabela,
            orient="vertical",
            command=self.tabela.yview
        )

        self.tabela.configure(
            yscrollcommand=scrollbar.set
        )

        self.tabela.pack(
            side="left",
            fill="both",
            expand=True
        )

        scrollbar.pack(
            side="right",
            fill="y"
        )

        # =====================================================
        # STATUS
        # =====================================================

        self.label_status = ttk.Label(
            root,
            text="Pronto.",
            anchor="center",
            style="Status.TLabel"
        )

        self.label_status.pack(
            pady=(0, 8)
        )

        # =====================================================
        # SAÍDAS
        # =====================================================

        frame_saida = ttk.LabelFrame(
            root,
            text="3. Arquivos de saída",
            padding=12
        )

        frame_saida.pack(
            fill="x",
            padx=20,
            pady=(0, 10)
        )

        frame_botoes_saida = ttk.Frame(
            frame_saida
        )

        frame_botoes_saida.pack()

        self.botao_excel = ttk.Button(
            frame_botoes_saida,
            text="Salvar Base Mestre Excel",
            command=self.salvar_excel,
            state="disabled",
            style="Secundario.TButton"
        )

        self.botao_excel.grid(
            row=0,
            column=0,
            padx=5
        )

        self.botao_csv = ttk.Button(
            frame_botoes_saida,
            text="Salvar Base Mestre CSV",
            command=self.salvar_csv,
            state="disabled",
            style="Secundario.TButton"
        )

        self.botao_csv.grid(
            row=0,
            column=1,
            padx=5
        )

        # =====================================================
        # RODAPÉ / ENCERRAR
        # =====================================================

        frame_rodape = ttk.Frame(
            root
        )

        frame_rodape.pack(
            fill="x",
            padx=20,
            pady=(0, 14)
        )

        rodape = ttk.Label(
            frame_rodape,
            text="Sistema de Cobrança — Módulo 1"
        )

        rodape.pack(
            side="left"
        )

        self.botao_encerrar = ttk.Button(
            frame_rodape,
            text="Encerrar sistema",
            command=self.encerrar_sistema,
            style="Secundario.TButton",
            width=18
        )

        self.botao_encerrar.pack(
            side="right",
            padx=(10, 0)
        )

        # Fecha pela bolinha vermelha no macOS
        # ou pelo X no Windows usando a mesma confirmação.
        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.encerrar_sistema
        )


    # =========================================================================
    # SELEÇÃO DOS PDFs
    # =========================================================================

    def selecionar_pdfs(self):

        arquivos = filedialog.askopenfilenames(
            title="Selecione os relatórios PDF",
            filetypes=[
                (
                    "Arquivos PDF",
                    "*.pdf"
                )
            ]
        )

        if not arquivos:
            return

        # =====================================================
        # LIMPAR EXECUÇÃO ANTERIOR
        # =====================================================

        # Limpar lista de arquivos selecionados
        for item in self.lista_arquivos.get_children():
            self.lista_arquivos.delete(item)

        # Limpar tabela de arquivos processados
        for item in self.tabela.get_children():
            self.tabela.delete(item)

        # Limpar dados da execução anterior
        self.base_mestre = None
        self.resumo = None
        self.validacao_estrutura = None
        self.relatorio = None

        # Zerar indicadores
        self.label_processados.config(
            text="PDFs processados: 0"
        )

        self.label_validos.config(
            text="PDFs validados: 0"
        )

        self.label_divergentes.config(
            text="Arquivos com problema: 0"
        )

        self.label_registros.config(
            text="Registros consolidados: 0"
        )

        # Desabilitar saídas até novo processamento
        self.botao_excel.config(
            state="disabled"
        )

        self.botao_csv.config(
            state="disabled"
        )

        self.arquivos_pdf = [
            Path(arquivo)
            for arquivo in arquivos
        ]

        quantidade = len(
            self.arquivos_pdf
        )

        self.label_selecionados.config(
            text=(
                f"{quantidade} "
                f"PDF(s) selecionado(s)."
            )
        )

        for item in self.lista_arquivos.get_children():
            self.lista_arquivos.delete(item)

        for arquivo in self.arquivos_pdf:

            self.lista_arquivos.insert(
                "",
                "end",
                values=(
                    arquivo.name,
                )
            )
        self.botao_processar.config(
            state="normal"
        )

        self.label_status.config(
            text=(
                "Nova seleção carregada. "
                "Clique em Processar arquivos."
            )
        )


    # =========================================================================
    # PROCESSAMENTO
    # =========================================================================

    def processar_arquivos(self):

        if not self.arquivos_pdf:
            return

        # -----------------------------------------------------
        # Limpar execução anterior
        # -----------------------------------------------------

        for item in self.tabela.get_children():

            self.tabela.delete(
                item
            )

        self.botao_processar.config(
            state="disabled"
        )

        self.botao_selecionar.config(
            state="disabled"
        )

        self.botao_excel.config(
            state="disabled"
        )

        self.botao_csv.config(
            state="disabled"
        )

        self.label_status.config(
            text="Processando..."
        )

        self.root.update()

        bases_validas = []

        registros_relatorio = []

        # -----------------------------------------------------
        # Processar PDF por PDF
        # -----------------------------------------------------

        for caminho_pdf in self.arquivos_pdf:

            self.label_status.config(
                text=(
                    f"Processando "
                    f"{caminho_pdf.name}..."
                )
            )

            self.root.update()

            try:

                resultado = processar_pdf(
                    caminho_pdf
                )

                valido = resultado_esta_ok(
                    resultado
                )

                dados = resultado[
                    "dados"
                ]

                cabecalho = resultado[
                    "cabecalho"
                ]

                status = (
                    "OK"
                    if valido
                    else "DIVERGENTE"
                )

                quantidade = len(
                    dados
                )

                registro_relatorio = {
                    "arquivo":
                        caminho_pdf.name,

                    "codigo_condominio":
                        cabecalho.get(
                            "codigo_condominio",
                            ""
                        ),

                    "condominio":
                        cabecalho.get(
                            "condominio",
                            ""
                        ),

                    "registros":
                        quantidade,

                    "status":
                        status
                }

                registros_relatorio.append(
                    registro_relatorio
                )

                if valido:

                    bases_validas.append(
                        dados
                    )

                self.tabela.insert(
                    "",
                    "end",
                    values=(
                        caminho_pdf.name,
                        registro_relatorio[
                            "codigo_condominio"
                        ],
                        registro_relatorio[
                            "condominio"
                        ],
                        quantidade,
                        status
                    )
                )

            except Exception as erro:

                registros_relatorio.append({
                    "arquivo":
                        caminho_pdf.name,

                    "codigo_condominio":
                        "",

                    "condominio":
                        "",

                    "registros":
                        0,

                    "status":
                        "ERRO"
                })

                self.tabela.insert(
                    "",
                    "end",
                    values=(
                        caminho_pdf.name,
                        "",
                        "",
                        0,
                        "ERRO"
                    )
                )

                print(
                    f"Erro em "
                    f"{caminho_pdf.name}: "
                    f"{erro}"
                )

        # -----------------------------------------------------
        # Relatório
        # -----------------------------------------------------

        self.relatorio = pd.DataFrame(
            registros_relatorio
        )

        # -----------------------------------------------------
        # Consolidar PDFs válidos
        # -----------------------------------------------------

        if bases_validas:

            self.base_mestre = pd.concat(
                bases_validas,
                ignore_index=True
            )

            self.resumo = gerar_resumo(
                self.base_mestre
            )

            self.validacao_estrutura = (
                validar_estrutura(
                    self.base_mestre
                )
            )

            self.botao_excel.config(
                state="normal"
            )

            self.botao_csv.config(
                state="normal"
            )

        else:

            self.base_mestre = None

            self.resumo = None

            self.validacao_estrutura = None

        # -----------------------------------------------------
        # Indicadores
        # -----------------------------------------------------

        total_processados = len(
            self.relatorio
        )

        total_validos = int(
            (
                self.relatorio[
                    "status"
                ]
                == "OK"
            ).sum()
        )

        total_divergentes = (
            total_processados
            -
            total_validos
        )

        total_registros = (
            len(self.base_mestre)
            if self.base_mestre
            is not None
            else 0
        )

        self.label_processados.config(
            text=(
                f"PDFs processados: "
                f"{total_processados}"
            )
        )

        self.label_validos.config(
            text=(
                f"PDFs validados: "
                f"{total_validos}"
            )
        )

        self.label_divergentes.config(
            text=(
                f"Arquivos com problema: "
                f"{total_divergentes}"
            )
        )

        self.label_registros.config(
            text=(
                f"Registros consolidados: "
                f"{total_registros}"
            )
        )

        # -----------------------------------------------------
        # Reativar controles
        # -----------------------------------------------------

        self.botao_selecionar.config(
            state="normal"
        )

        self.botao_processar.config(
            state="normal"
        )

        # -----------------------------------------------------
        # Status final
        # -----------------------------------------------------

        if total_validos == total_processados:

            self.label_status.config(
                text=(
                    "Processamento concluído. "
                    "Todos os PDFs foram validados."
                )
            )

        else:

            self.label_status.config(
                text=(
                    "Processamento concluído. "
                    "Existem arquivos com erro "
                    "ou divergência."
                )
            )


    # =========================================================================
    # SALVAR EXCEL
    # =========================================================================

    def salvar_excel(self):

        if self.base_mestre is None:
            return

        arquivo = filedialog.asksaveasfilename(
            title="Salvar Base Mestre",
            defaultextension=".xlsx",
            initialfile="base_mestre.xlsx",
            filetypes=[
                (
                    "Arquivo Excel",
                    "*.xlsx"
                )
            ]
        )

        if not arquivo:
            return

        arquivo = Path(
            arquivo
        )

        try:

            with pd.ExcelWriter(
                arquivo,
                engine="openpyxl"
            ) as writer:

                self.base_mestre.to_excel(
                    writer,
                    sheet_name="Dados",
                    index=False
                )

                self.resumo.to_excel(
                    writer,
                    sheet_name="Resumo",
                    index=False
                )

                self.validacao_estrutura.to_excel(
                    writer,
                    sheet_name="Validação",
                    index=False
                )

            formatar_excel(
                arquivo
            )

            messagebox.showinfo(
                "Sistema de Cobrança",
                "Base Mestre Excel gerada com sucesso."
            )

        except Exception as erro:

            messagebox.showerror(
                "Erro",
                (
                    "Não foi possível salvar "
                    "o arquivo Excel.\n\n"
                    f"{erro}"
                )
            )


    # =========================================================================
    # SALVAR CSV
    # =========================================================================

    def salvar_csv(self):

        if self.base_mestre is None:
            return

        arquivo = filedialog.asksaveasfilename(
            title="Salvar Base Mestre CSV",
            defaultextension=".csv",
            initialfile="base_mestre.csv",
            filetypes=[
                (
                    "Arquivo CSV",
                    "*.csv"
                )
            ]
        )

        if not arquivo:
            return

        try:

            self.base_mestre.to_csv(
                arquivo,
                index=False,
                sep=";",
                encoding="utf-8-sig",
                decimal=","
            )

            messagebox.showinfo(
                "Sistema de Cobrança",
                "Base Mestre CSV gerada com sucesso."
            )

        except Exception as erro:

            messagebox.showerror(
                "Erro",
                (
                    "Não foi possível salvar "
                    "o arquivo CSV.\n\n"
                    f"{erro}"
                )
            )


    # =========================================================================
    # ENCERRAR SISTEMA
    # =========================================================================

    def encerrar_sistema(self):

        resposta = messagebox.askyesno(
            "Encerrar sistema",
            (
                "Deseja realmente encerrar "
                "o Sistema de Cobrança?"
            )
        )

        if resposta:
            self.root.destroy()


# =============================================================================
# INICIAR APLICAÇÃO
# =============================================================================

def main():

    root = tk.Tk()

    SistemaCobrancaApp(
        root
    )

    root.mainloop()


if __name__ == "__main__":
    main()