# =============================================================================
# SISTEMA DE COBRANÇA - MÓDULO 1
# Conversão de Relatório de Inadimplência PDF para Base Mestre
# =============================================================================
#
# OBJETIVO
# -------
# Este programa lê um relatório de inadimplência em PDF, extrai os dados
# dos débitos e os transforma em uma tabela estruturada.
#
# Nesta versão, o programa:
#
#   1. Lê o arquivo PDF;
#   2. Extrai os dados do condomínio;
#   3. Identifica cada economia/devedor;
#   4. Extrai os lançamentos em aberto;
#   5. Separa o valor corrigido do "Nosso Número";
#   6. Converte os valores financeiros para formato numérico;
#   7. Soma os valores extraídos;
#   8. Compara as somas com os totais informados no próprio PDF;
#   9. Gera um arquivo CSV estruturado.
#
#
# ESTRUTURA ESPERADA DE PASTAS
# ----------------------------
#
# modulo1/
# │
# ├── .venv/
# │
# ├── parser_pdf_base_mestre_v1.py
# │
# ├── pdfs/
# │   └── 153.pdf
# │
# └── saida/
#
#
# PREPARAÇÃO DO AMBIENTE
# ----------------------
#
# Abrir o Terminal na pasta "modulo1".
#
# Ativar o ambiente virtual:
#
#     source .venv/bin/activate
#
# O Terminal deverá passar a mostra o prompt assim:
#
#     (.venv)
#
#
# DEPENDÊNCIAS
# ------------
#
# Caso seja a primeira execução ou o ambiente virtual tenha sido recriado:
#
#     python -m pip install pdfplumber pandas openpyxl
#
#
# ARQUIVO DE ENTRADA
# ------------------
#
# Colocar o relatório PDF dentro da pasta:
#
#     pdfs/
#
# Nesta versão de teste, o programa está configurado para processar:
#
#     pdfs/153.pdf
#
# O arquivo utilizado pode ser alterado na seção:
#
#     if __name__ == "__main__":
#
#
# EXECUÇÃO
# --------
#
# No Terminal, estando dentro da pasta "modulo1" e com o .venv ativado:
#
#     python parser_pdf_base_mestre_v1.py
#
#
# RESULTADO
# ---------
#
# Durante a execução, o programa apresenta no Terminal:
#
#   - nome do arquivo processado;
#   - quantidade de registros encontrados;
#   - totais financeiros extraídos;
#   - comparação entre os totais extraídos e os totais do PDF;
#   - indicação "OK" ou "DIVERGENTE" para cada total.
#
# Se o processamento for concluído, será criado:
#
#     saida/base_mestre_153.csv
#
#
# VALIDAÇÃO
# ---------
#
# O arquivo não deve ser considerado corretamente convertido apenas porque
# os registros foram encontrados.
#
# Os seguintes totais calculados a partir dos registros devem coincidir com
# os totais apresentados no relatório original:
#
#   - Valor Base
#   - Multa
#   - Juros
#   - Correção
#   - Valor Corrigido
#
# Uma divergência indica que alguma linha do PDF pode não ter sido
# interpretada corretamente e deve ser investigada antes de utilizar
# os dados na Base Mestre.
#
#
# VERSÃO
# ------
#
# V1 - Parser inicialmente desenvolvido e validado com o relatório 153.pdf.
#
# Resultado da validação inicial:
#
#   Registros:       16
#   Valor Base:      R$ 7.113,03
#   Multa:           R$   142,27
#   Juros:           R$   136,65
#   Correção:        R$   100,60
#   Valor Corrigido: R$ 7.492,55
#
# Próxima etapa:
# testar o parser com relatórios de outros condomínios e adaptar as regras
# somente quando forem encontradas diferenças reais no layout dos PDFs.
#
# V1.1
# - Parser validado inicialmente com 153.pdf.
# - Adicionado suporte ao campo Advogado opcional no final da linha.
# - 234.pdf passou a ser reconhecido integralmente: 22 registros.
#
# V1.2
# - Processamento em lote de todos os PDFs da pasta /pdfs.
# - Validação financeira individual de cada relatório.
# - PDFs divergentes não são incorporados à Base Mestre.
# - Suporte ao campo Advogado opcional.
# - Suporte a diferentes formatos de economia:
#     424
#     304*
#     08 - 4 (01)
#     SALÃO 1 (01)
#     1E (01)
# - Geração da Base Mestre consolidada.
# - Geração das abas Dados e Resumo no Excel.
#
# ============================================================
# SISTEMA DE COBRANÇA - MÓDULO 1
# Parser de Relatórios de Inadimplência
#
# Versão: 1.3
# Status: ESTÁVEL
#
# Funcionalidades:
# - Leitura dos relatórios PDF
# - Processamento em lote
# - Extração dos dados dos lançamentos
# - Suporte a diferentes formatos de economia
# - Suporte a advogado opcional
# - Validação dos totais financeiros por PDF
# - Exclusão de PDFs divergentes da Base Mestre
# - Geração da Base Mestre
# - Geração do Resumo por economia/devedor
# - Validação estrutural
# - Exportação para CSV e Excel
# - Formatação automática do Excel
# ============================================================
#
# =============================================================================

from pathlib import Path
import re

import pdfplumber
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURAÇÕES
# ============================================================

COLUNAS_VALORES = [
    "vlr_base",
    "multa",
    "juros",
    "correcao",
    "vlr_corrigido"
]


COLUNAS_MESTRE = [
    "codigo_condominio",
    "condominio",
    "economia",
    "nome_condomino",
    "competencia",
    "data_vencimento",
    "tipo",
    "vlr_base",
    "multa",
    "juros",
    "correcao",
    "vlr_corrigido",
    "assessor",
    "imobiliaria",
    "endereco",
    "nosso_numero",
    "advogado"
]


# ============================================================
# CABEÇALHO DO PDF
# ============================================================

def extrair_cabecalho(texto):
    """
    Extrai:

    codigo_condominio
    condominio
    assessor
    imobiliaria
    endereco
    """

    linhas = texto.splitlines()

    if len(linhas) < 2:
        return None

    linha_cabecalho = linhas[0].strip()
    endereco = linhas[1].strip()

    padrao = re.compile(
        r"^Condomínio:\s*"
        r"(?P<codigo>\d+)-"
        r"(?P<condominio>.+?)\s+"
        r"\(Assessor:\s*(?P<assessor>[^)]+)\)\s+"
        r"Imobiliária\s+(?P<imobiliaria>.+)$"
    )

    m = padrao.match(linha_cabecalho)

    if not m:
        return None

    return {
        "codigo_condominio": m.group("codigo"),
        "condominio": m.group("condominio").strip(),
        "assessor": m.group("assessor").strip(),
        "imobiliaria": m.group("imobiliaria").strip(),
        "endereco": endereco
    }


# ============================================================
# CAMPO FINAL:
# VALOR CORRIGIDO + NOSSO NÚMERO
# ============================================================

def separar_valor_corrigido_nosso_numero(campo):
    """
    Separa:

    valor corrigido
    nosso número
    advogado (quando existir)

    Exemplos:

    431,370000026492705

    retorna:
        vlr_corrigido = 431,37
        nosso_numero  = 0000026492705
        advogado      = ""

    --------------------------------------------------

    457,380000025433789 DENISE DE CASSIA BAIOTO EBBESEN

    retorna:
        vlr_corrigido = 457,38
        nosso_numero  = 0000025433789
        advogado      = DENISE DE CASSIA BAIOTO EBBESEN
    """

    padrao = re.compile(
        r"^(?P<vlr_corrigido>\d{1,3}(?:\.\d{3})*,\d{2})"
        r"(?P<nosso_numero>\d+)"
        r"(?:\s+(?P<advogado>.+))?$"
    )

    m = padrao.match(campo.strip())

    if not m:
        return None

    return {
        "vlr_corrigido": m.group("vlr_corrigido"),
        "nosso_numero": m.group("nosso_numero"),
        "advogado": (
            m.group("advogado").strip()
            if m.group("advogado")
            else ""
        )
    }

# ============================================================
# LINHA QUE INICIA UMA NOVA ECONOMIA / DEVEDOR
# ============================================================

def interpretar_linha_completa(linha):
    """
    Exemplo:

    424 ALEXANDRE PETERS REGO
    07/2026 08/07/2026 N
    422,91 8,46 0,00 0,00
    431,370000026492705
    """

    padrao = re.compile(
        r"^(?P<economia>"
            r".+?\(\d+\)"          # SALÃO 1 (01), 1E (01), 08 - 4 (01)
            r"|\d+\s+BOX\s+\d+"        # 205 BOX 05, 304 BOX 03
            r"|\d+\s*-\s*\d+"      # 2 - 1248
            r"|\S+"                 # 424, 304*, 1E etc.
        r")\s+"
        r"(?P<nome>.+?)\s+"
        r"(?P<competencia>\d{2}/\d{4})\s+"
        r"(?P<vencimento>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<tipo>[NAE])\s+"
        r"(?P<vlr_base>[\d.,]+)\s+"
        r"(?P<multa>[\d.,]+)\s+"
        r"(?P<juros>[\d.,]+)\s+"
        r"(?P<correcao>[\d.,]+)\s+"
        r"(?P<final>.+)$",
        re.IGNORECASE
    )

    m = padrao.match(linha.strip())

    if not m:
        return None

    final = separar_valor_corrigido_nosso_numero(
        m.group("final")
    )

    if final is None:
        return None

    return {
        "economia": m.group("economia"),
        "nome_condomino": m.group("nome"),
        "competencia": m.group("competencia"),
        "data_vencimento": m.group("vencimento"),
        "tipo": m.group("tipo"),
        "vlr_base": m.group("vlr_base"),
        "multa": m.group("multa"),
        "juros": m.group("juros"),
        "correcao": m.group("correcao"),
        "vlr_corrigido": final["vlr_corrigido"],
        "nosso_numero": final["nosso_numero"],
        "advogado": final["advogado"]
    }


# ============================================================
# LINHA DE CONTINUAÇÃO DA MESMA ECONOMIA
# ============================================================

def interpretar_linha_continuacao(
    linha,
    economia,
    nome_condomino
):
    """
    Exemplo:

    07/2026 08/07/2026 N
    451,84 9,04 0,00 0,00
    460,880000026492708
    """

    padrao = re.compile(
        r"^(?P<competencia>\d{2}/\d{4})\s+"
        r"(?P<vencimento>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<tipo>[NAE])\s+"
        r"(?P<vlr_base>[\d.,]+)\s+"
        r"(?P<multa>[\d.,]+)\s+"
        r"(?P<juros>[\d.,]+)\s+"
        r"(?P<correcao>[\d.,]+)\s+"
        r"(?P<final>.+)$"
    )

    m = padrao.match(linha.strip())

    if not m:
        return None

    final = separar_valor_corrigido_nosso_numero(
        m.group("final")
    )

    if final is None:
        return None

    return {
        "economia": economia,
        "nome_condomino": nome_condomino,
        "competencia": m.group("competencia"),
        "data_vencimento": m.group("vencimento"),
        "tipo": m.group("tipo"),
        "vlr_base": m.group("vlr_base"),
        "multa": m.group("multa"),
        "juros": m.group("juros"),
        "correcao": m.group("correcao"),
        "vlr_corrigido": final["vlr_corrigido"],
        "nosso_numero": final["nosso_numero"],
        "advogado": final["advogado"]

    }


# ============================================================
# TOTAIS INFORMADOS NO RODAPÉ DO PDF
# ============================================================

def extrair_totais_pdf(texto):
    """
    Extrai os totais impressos no relatório.
    """

    totais = {}

    padroes = {
        "vlr_base":
            r"Vlr\.Base:\s*([\d\.,]+)",

        "multa":
            r"Multa:\s*([\d\.,]+)",

        "juros":
            r"Juros:\s*([\d\.,]+)",

        "correcao":
            r"Corr:\s*([\d\.,]+)",

        "vlr_corrigido":
            r"Vlr\.Corrigido:\s*([\d\.,]+)"
    }

    for campo, padrao in padroes.items():

        m = re.search(
            padrao,
            texto
        )

        if m:

            valor = (
                m.group(1)
                .replace(".", "")
                .replace(",", ".")
            )

            totais[campo] = float(valor)

    return totais


# ============================================================
# LEITURA DO PDF
# ============================================================

def ler_pdf(caminho_pdf):
    """
    Extrai o texto de todas as páginas.
    """

    textos_paginas = []

    with pdfplumber.open(caminho_pdf) as pdf:

        for pagina in pdf.pages:

            texto = (
                pagina.extract_text()
                or ""
            )

            textos_paginas.append(
                texto
            )

    return "\n".join(
        textos_paginas
    )


# ============================================================
# PROCESSAMENTO COMPLETO DE UM PDF
# ============================================================

def processar_pdf(caminho_pdf):

    caminho_pdf = Path(
        caminho_pdf
    )

    print(
        f"Processando: {caminho_pdf.name}"
    )

    # --------------------------------------------------------
    # 1. Ler PDF
    # --------------------------------------------------------

    texto_completo = ler_pdf(
        caminho_pdf
    )

    # --------------------------------------------------------
    # 2. Cabeçalho
    # --------------------------------------------------------

    cabecalho = extrair_cabecalho(
        texto_completo
    )

    if cabecalho is None:

        raise ValueError(
            "Não foi possível identificar "
            f"o cabeçalho de {caminho_pdf.name}"
        )

    # --------------------------------------------------------
    # 3. Interpretar lançamentos
    # --------------------------------------------------------

    registros = []

    economia_atual = None
    nome_atual = None

    for linha in texto_completo.splitlines():

        # Tenta primeiro como início
        # de nova economia

        registro = interpretar_linha_completa(
            linha
        )

        if registro is not None:

            economia_atual = (
                registro["economia"]
            )

            nome_atual = (
                registro[
                    "nome_condomino"
                ]
            )

            registros.append(
                registro
            )

            continue

        # Caso contrário,
        # tenta como continuação

        if economia_atual is not None:

            registro = (
                interpretar_linha_continuacao(
                    linha,
                    economia_atual,
                    nome_atual
                )
            )

            if registro is not None:

                registros.append(
                    registro
                )

    # --------------------------------------------------------
    # 4. Criar DataFrame
    # --------------------------------------------------------

    df = pd.DataFrame(
        registros
    )

    if df.empty:

        raise ValueError(
            f"Nenhum lançamento encontrado "
            f"em {caminho_pdf.name}"
        )

    # --------------------------------------------------------
    # 5. Adicionar dados do cabeçalho
    # --------------------------------------------------------

    df["codigo_condominio"] = (
        cabecalho[
            "codigo_condominio"
        ]
    )

    df["condominio"] = (
        cabecalho[
            "condominio"
        ]
    )

    df["assessor"] = (
        cabecalho[
            "assessor"
        ]
    )

    df["imobiliaria"] = (
        cabecalho[
            "imobiliaria"
        ]
    )

    df["endereco"] = (
        cabecalho[
            "endereco"
        ]
    )

    # --------------------------------------------------------
    # 6. Converter valores para número
    # --------------------------------------------------------

    for coluna in COLUNAS_VALORES:

        df[coluna] = (
            df[coluna]
            .str.replace(
                ".",
                "",
                regex=False
            )
            .str.replace(
                ",",
                ".",
                regex=False
            )
            .astype(float)
        )

    # --------------------------------------------------------
    # 7. Reordenar colunas
    # --------------------------------------------------------

    df = df[
        COLUNAS_MESTRE
    ]

    # --------------------------------------------------------
    # 8. Totais do PDF
    # --------------------------------------------------------

    totais_pdf = (
        extrair_totais_pdf(
            texto_completo
        )
    )

    # --------------------------------------------------------
    # 9. Totais calculados
    # --------------------------------------------------------

    totais_extraidos = {

        campo:
            round(
                df[campo].sum(),
                2
            )

        for campo
        in COLUNAS_VALORES
    }

    # --------------------------------------------------------
    # 10. Validação
    # --------------------------------------------------------

    validacao = {}

    for campo in COLUNAS_VALORES:

        esperado = (
            totais_pdf.get(
                campo
            )
        )

        encontrado = (
            totais_extraidos[
                campo
            ]
        )

        ok = (
            esperado is not None
            and
            abs(
                esperado
                -
                encontrado
            )
            < 0.01
        )

        validacao[campo] = {
            "pdf":
                esperado,

            "extraido":
                encontrado,

            "ok":
                ok
        }

    return {
        "arquivo":
            caminho_pdf.name,

        "cabecalho":
            cabecalho,

        "dados":
            df,

        "totais_pdf":
            totais_pdf,

        "totais_extraidos":
            totais_extraidos,

        "validacao":
            validacao
    }


# ============================================================
# MOSTRAR RESULTADO DA VALIDAÇÃO
# ============================================================

def mostrar_validacao(resultado):

    print()
    print(
        "VALIDAÇÃO FINANCEIRA"
    )

    print(
        "-" * 65
    )

    for campo, valores in (
        resultado[
            "validacao"
        ].items()
    ):

        status = (
            "OK"
            if valores["ok"]
            else "DIVERGENTE"
        )

        pdf = valores["pdf"]
        extraido = (
            valores["extraido"]
        )

        print(
            f"{campo:15} "
            f"PDF={pdf:10.2f} "
            f"Extraído={extraido:10.2f} "
            f"{status}"
        )

    print(
        "-" * 65
    )

    print(
        "Quantidade de registros:",
        len(
            resultado[
                "dados"
            ]
        )
    )


# ============================================================
# EXPORTAR CSV
# ============================================================

def exportar_csv(
    resultado,
    caminho_saida
):

    df = resultado[
        "dados"
    ]

    df.to_csv(
        caminho_saida,
        index=False,
        sep=";",
        encoding="utf-8-sig",
        decimal=","
    )

# ============================================================
# GERAR RESUMO POR DEVEDOR
# ============================================================

def gerar_resumo(base_mestre):

    resumo = (
        base_mestre
        .groupby(
            [
                "codigo_condominio",
                "condominio",
                "economia",
                "nome_condomino"
            ],
            dropna=False,
            as_index=False
        )
        .agg(
            valor_total_devido_base=(
                "vlr_base",
                "sum"
            ),
            valor_total_corrigido=(
                "vlr_corrigido",
                "sum"
            )
        )
    )

    return resumo

# ============================================================
# VALIDAR QUALIDADE ESTRUTURAL DOS DADOS
# ============================================================

def validar_estrutura(base_mestre):

    alertas = []

    # --------------------------------------------------------
    # 1. Economia vazia
    # --------------------------------------------------------

    for indice, registro in base_mestre.iterrows():

        economia = str(registro["economia"]).strip()
        nome = str(registro["nome_condomino"]).strip()
        nosso_numero = str(registro["nosso_numero"]).strip()

        if (
            not economia
            or economia.lower() == "nan"
        ):
            alertas.append({
                "tipo_alerta": "Economia vazia",
                "codigo_condominio": registro["codigo_condominio"],
                "condominio": registro["condominio"],
                "economia": registro["economia"],
                "nome_condomino": registro["nome_condomino"],
                "nosso_numero": registro["nosso_numero"]
            })

        # ----------------------------------------------------
        # 2. Nome do condômino vazio
        # ----------------------------------------------------

        if (
            not nome
            or nome.lower() == "nan"
        ):
            alertas.append({
                "tipo_alerta": "Nome do condômino vazio",
                "codigo_condominio": registro["codigo_condominio"],
                "condominio": registro["condominio"],
                "economia": registro["economia"],
                "nome_condomino": registro["nome_condomino"],
                "nosso_numero": registro["nosso_numero"]
            })

        # ----------------------------------------------------
        # 3. Nome começando com BOX
        #
        # Pode indicar que parte da economia foi
        # incorretamente incorporada ao nome.
        # ----------------------------------------------------

        if re.match(
            r"^BOX\s+\d+",
            nome,
            re.IGNORECASE
        ):
            alertas.append({
                "tipo_alerta": "Nome inicia com BOX",
                "codigo_condominio": registro["codigo_condominio"],
                "condominio": registro["condominio"],
                "economia": registro["economia"],
                "nome_condomino": registro["nome_condomino"],
                "nosso_numero": registro["nosso_numero"]
            })

        # ----------------------------------------------------
        # 4. Nosso Número vazio
        # ----------------------------------------------------

        if (
            not nosso_numero
            or nosso_numero.lower() == "nan"
        ):
            alertas.append({
                "tipo_alerta": "Nosso Número vazio",
                "codigo_condominio": registro["codigo_condominio"],
                "condominio": registro["condominio"],
                "economia": registro["economia"],
                "nome_condomino": registro["nome_condomino"],
                "nosso_numero": registro["nosso_numero"]
            })

    # --------------------------------------------------------
    # 5. Nosso Número duplicado
    # --------------------------------------------------------

    duplicados = base_mestre[
        base_mestre["nosso_numero"].duplicated(
            keep=False
        )
    ]

    for _, registro in duplicados.iterrows():

        alertas.append({
            "tipo_alerta": "Nosso Número duplicado",
            "codigo_condominio": registro["codigo_condominio"],
            "condominio": registro["condominio"],
            "economia": registro["economia"],
            "nome_condomino": registro["nome_condomino"],
            "nosso_numero": registro["nosso_numero"]
        })

    # --------------------------------------------------------
    # Criar DataFrame de validação
    # --------------------------------------------------------

    colunas = [
        "tipo_alerta",
        "codigo_condominio",
        "condominio",
        "economia",
        "nome_condomino",
        "nosso_numero"
    ]

    return pd.DataFrame(
        alertas,
        columns=colunas
    )

# ============================================================
# PROGRAMA PRINCIPAL
# ============================================================

if __name__ == "__main__":

    pasta_pdfs = Path("pdfs")
    pasta_saida = Path("saida")

    pasta_saida.mkdir(exist_ok=True)

    arquivos_pdf = sorted(
        pasta_pdfs.glob("*.pdf")
    )

    if not arquivos_pdf:
        raise ValueError(
            "Nenhum PDF encontrado na pasta pdfs."
        )

    bases_validas = []
    relatorio = []

    for arquivo_pdf in arquivos_pdf:

        print()
        print("=" * 70)
        print(f"Processando: {arquivo_pdf.name}")
        print("=" * 70)

        try:

            resultado = processar_pdf(
                arquivo_pdf
            )

            mostrar_validacao(
                resultado
            )

            valido = all(
                item["ok"]
                for item
                in resultado["validacao"].values()
            )

            quantidade = len(
                resultado["dados"]
            )

            if valido:

                print(
                    f"{arquivo_pdf.name}: VALIDADO"
                )

                bases_validas.append(
                    resultado["dados"]
                )

                status = "OK"

            else:

                print(
                    f"{arquivo_pdf.name}: DIVERGENTE"
                )

                print(
                    "Este PDF não será incluído "
                    "na Base Mestre."
                )

                status = "DIVERGENTE"

            relatorio.append({
                "arquivo": arquivo_pdf.name,
                "codigo_condominio":
                    resultado["cabecalho"]["codigo_condominio"],
                "condominio":
                    resultado["cabecalho"]["condominio"],
                "registros":
                    quantidade,
                "status":
                    status
            })

        except Exception as erro:

            print(
                f"ERRO em {arquivo_pdf.name}: {erro}"
            )

            relatorio.append({
                "arquivo":
                    arquivo_pdf.name,
                "codigo_condominio":
                    "",
                "condominio":
                    "",
                "registros":
                    0,
                "status":
                    f"ERRO: {erro}"
            })

    # Junta todos os PDFs validados
    if bases_validas:

        base_mestre = pd.concat(
            bases_validas,
            ignore_index=True
        )

    else:

        base_mestre = pd.DataFrame(
            columns=COLUNAS_MESTRE
        )

    # Cria o relatório do processamento
    relatorio_df = pd.DataFrame(
        relatorio
    )

# ============================================================
# EXPORTAR BASE MESTRE EM CSV
# ============================================================

    arquivo_csv = (
        pasta_saida
        /
        "base_mestre.csv"
    )

    base_mestre.to_csv(
        arquivo_csv,
        index=False,
        sep=";",
        encoding="utf-8-sig",
        decimal=","
    )


# ============================================================
# GERAR RESUMO POR DEVEDOR
# ============================================================

    resumo = gerar_resumo(base_mestre)

    validacao_estrutura = validar_estrutura(
        base_mestre
    )

    linha_total = pd.DataFrame([{
        "codigo_condominio": "",
        "condominio": "TOTAL GERAL",
        "economia": "",
        "nome_condomino": "",
        "valor_total_devido_base": resumo["valor_total_devido_base"].sum(),
        "valor_total_corrigido": resumo["valor_total_corrigido"].sum()
}])

    resumo = pd.concat(
        [resumo, linha_total],
        ignore_index=True
    )

# ============================================================
# EXPORTAR EXCEL COM DUAS ABAS:
# Dados + Resumo
# ============================================================

    arquivo_excel = (
        pasta_saida
        /
        "base_mestre.xlsx"
    )

    with pd.ExcelWriter(
        arquivo_excel,
        engine="openpyxl"
    ) as writer:

        base_mestre.to_excel(
            writer,
            sheet_name="Dados",
            index=False
        )

        resumo.to_excel(
            writer,
            sheet_name="Resumo",
            index=False
        )

        validacao_estrutura.to_excel(
            writer,
            sheet_name="Validação",
            index=False
        )

    # ============================================================
    # FORMATAR O ARQUIVO EXCEL
    # ============================================================

    wb = load_workbook(arquivo_excel)

    cor_cabecalho = "4F81BD"

    # Borda fina para todas as células
    borda_fina = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7")
    )

    for nome_aba in [
        "Dados",
        "Resumo",
        "Validação"
    ]:
        ws = wb[nome_aba]

        # Aplicar bordas em todas as células utilizadas
        for linha in ws.iter_rows():
            for cell in linha:
                cell.border = borda_fina

        # Formatar cabeçalho
        for cell in ws[1]:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=cor_cabecalho
            )

            cell.font = Font(
                color="FFFFFF",
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # Congelar primeira linha
        ws.freeze_panes = "A2"

        # Colocar filtro
        ws.auto_filter.ref = ws.dimensions

        # Ajustar largura das colunas
        for coluna in ws.columns:

            maior = 0

            letra = get_column_letter(
                coluna[0].column
            )

            for cell in coluna:

                valor = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                if len(valor) > maior:
                    maior = len(valor)

            ws.column_dimensions[letra].width = min(
                maior + 2,
                40
            )


    # ============================================================
    # FORMATAR VALORES MONETÁRIOS - ABA DADOS
    # ============================================================

    ws_dados = wb["Dados"]

    cabecalhos_dados = {
        cell.value: cell.column
        for cell in ws_dados[1]
    }

    for nome_coluna in [
        "vlr_base",
        "multa",
        "juros",
        "correcao",
        "vlr_corrigido"
    ]:

        if nome_coluna in cabecalhos_dados:

            col = cabecalhos_dados[nome_coluna]

            for row in range(
                2,
                ws_dados.max_row + 1
            ):

                ws_dados.cell(
                    row=row,
                    column=col
                ).number_format = '#,##0.00'


    # ============================================================
    # FORMATAR VALORES MONETÁRIOS - ABA RESUMO
    # ============================================================

    ws_resumo = wb["Resumo"]

    cabecalhos_resumo = {
        cell.value: cell.column
        for cell in ws_resumo[1]
    }

    for nome_coluna in [
        "valor_total_devido_base",
        "valor_total_corrigido"
    ]:

        if nome_coluna in cabecalhos_resumo:

            col = cabecalhos_resumo[nome_coluna]

            for row in range(
                2,
                ws_resumo.max_row + 1
            ):

                ws_resumo.cell(
                    row=row,
                    column=col
                ).number_format = '#,##0.00'


    # ============================================================
    # SALVAR A FORMATAÇÃO
    # ============================================================

    wb.save(arquivo_excel)

# ============================================================
# EXPORTAR RELATÓRIO DE PROCESSAMENTO
# ============================================================

    arquivo_relatorio = (
        pasta_saida
        /
        "relatorio_processamento.xlsx"
    )

    relatorio_df.to_excel(
        arquivo_relatorio,
        index=False
    )

    print()
    print("=" * 70)
    print("PROCESSAMENTO FINALIZADO")
    print("=" * 70)

    print(
        f"PDFs processados: "
        f"{len(relatorio_df)}"
    )

    print(
        f"PDFs válidos: "
        f"{(relatorio_df['status'] == 'OK').sum()}"
    )

    print(
        f"Registros na Base Mestre: "
        f"{len(base_mestre)}"
    )

    print()
    print(
        f"Base Mestre: "
        f"{arquivo_csv}"
    )

    print(
        f"Relatório: "
        f"{arquivo_relatorio}"
    )

    print(
        f"Excel Mestre: "
        f"{arquivo_excel}"
    )

    print(
        f"Alertas estruturais: "
        f"{len(validacao_estrutura)}"
    )